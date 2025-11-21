import pandas as pd
from datetime import datetime
import os
import tkinter as tk
from tkinter import filedialog


class ExportManager:
    @staticmethod
    def export_to_excel(data, headers, filename_prefix, sheet_name="Datos"):
        """
        Exporta datos a Excel con formato profesional

        Args:
            data: Lista de tuplas con los datos
            headers: Lista de nombres de columnas
            filename_prefix: Prefijo para el nombre del archivo
            sheet_name: Nombre de la hoja de Excel

        Returns:
            tuple: (nombre_archivo, error) - error es None si fue exitoso
        """
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"{filename_prefix}_{timestamp}.xlsx"
            
            # Mostrar diálogo para seleccionar ubicación
            filename = ExportManager._get_save_filename(default_filename)
            
            if not filename:  # Usuario canceló la operación
                return None, "Exportación cancelada por el usuario"

            # Crear DataFrame
            df = pd.DataFrame(data, columns=headers)

            # Crear writer de Excel
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Obtener la hoja de trabajo
                worksheet = writer.sheets[sheet_name]

                # Aplicar formato profesional
                ExportManager._apply_excel_formatting(worksheet, df)

            # Verificar que el archivo se creó correctamente
            if os.path.exists(filename):
                return filename, None
            else:
                return None, "No se pudo crear el archivo"

        except Exception as e:
            return None, str(e)

    @staticmethod
    def _get_save_filename(default_filename):
        """
        Muestra diálogo para seleccionar ubicación y nombre del archivo
        
        Args:
            default_filename: Nombre por defecto del archivo
            
        Returns:
            str: Ruta completa del archivo seleccionada o None si se cancela
        """
        # Crear ventana temporal (no visible)
        root = tk.Tk()
        root.withdraw()  # Ocultar ventana principal
        
        # Mostrar diálogo de guardado
        filename = filedialog.asksaveasfilename(
            title="Guardar archivo Excel",
            defaultextension=".xlsx",
            filetypes=[
                ("Archivos Excel", "*.xlsx"),
                ("Todos los archivos", "*.*")
            ],
            initialfile=default_filename
        )
        
        # Cerrar ventana temporal
        root.destroy()
        
        return filename if filename else None

    @staticmethod
    def _apply_excel_formatting(worksheet, df):
        """
        Aplica formato profesional a la hoja de Excel
        """
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid")
        border_style = Side(border_style="thin", color="000000")
        border = Border(left=border_style, right=border_style,
                        top=border_style, bottom=border_style)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # Aplicar formato a los encabezados
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align

        # Aplicar formato a los datos
        # +2 porque row1 es encabezado y Excel base 1
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = left_align

        # Ajustar ancho de columnas automáticamente
        for col_num, column_title in enumerate(df.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)

            # Longitud del encabezado
            max_length = max(max_length, len(str(column_title)))

            # Longitud de los datos en la columna
            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Ajustar ancho con un poco de margen
            worksheet.column_dimensions[column_letter].width = max_length + 4

        # Congelar paneles (fijar encabezados)
        worksheet.freeze_panes = "A2"

        # Agregar filtros automáticos
        worksheet.auto_filter.ref = worksheet.dimensions

    @staticmethod
    def export_inventory(data):
        """
        Exportación específica para inventario
        """
        headers = ["Nro", "Producto", "Marca", "Categoría",
                "Código", "Stock", "Stock mínimo", "Ubicación", "Estado"]
        return ExportManager.export_to_excel(data, headers, "inventario", "Inventario")
    
    @staticmethod
    def export_movements(data):
        """
        Exportación específica para movimientos
        """
        headers = ["Nro", "Fecha", "Tipo", "Producto",
                   "Cantidad", "Ubicación", "Responsable", "Referencia"]
        return ExportManager.export_to_excel(data, headers, "movimientos", "Movimientos")

    @staticmethod
    def export_purchases(data):
        """
        Exportación específica para compras
        """
        headers = ["Nro", "Producto", "Cantidad", "Motivo",
                   "Prioridad", "Proveedor", "Fecha", "Estado"]
        return ExportManager.export_to_excel(data, headers, "solicitudes_compras", "Solicitudes_Compra")

    @staticmethod
    def export_suppliers(data):
        """
        Exportación específica para proveedores
        """
        headers = ["Nro", "Nombre", "Contacto", "Teléfono", "Email",
                   "Valoración", "Manejo de Precios", "Categorías"]
        return ExportManager.export_to_excel(data, headers, "proveedores", "Proveedores")

    @staticmethod
    def export_requests(data):
        """
        Exportación específica para solicitudes
        """
        headers = ["Nro", "Fecha", "Departamento",
                   "Solicitante", "Referencia", "Responsable Entrega"]
        return ExportManager.export_to_excel(data, headers, "solicitudes", "Solicitudes")

    @staticmethod
    def export_with_custom_format(data, headers, filename_prefix, sheet_name="Datos"):
        """
        Método para exportación personalizada con formato específico
        """
        return ExportManager.export_to_excel(data, headers, filename_prefix, sheet_name)